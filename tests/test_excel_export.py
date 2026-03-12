from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from container_planner.excel_export import build_excel_report
from container_planner.reporting import build_container_kpi_rows


def test_build_excel_report_creates_required_sheets():
    placements_df = pd.DataFrame(
        [
            {
                "container_label": "40HC ①",
                "container_type": "40HC",
                "container_index": 1,
                "cargo_piece_id": "A-1",
                "weight_kg": 120,
                "m3": 1.2,
                "placed_x_cm": 0,
                "placed_y_cm": 0,
                "orient_L_cm": 100,
                "orient_W_cm": 80,
            },
            {
                "container_label": "40HC ①",
                "container_type": "40HC",
                "container_index": 1,
                "cargo_piece_id": "B-1",
                "weight_kg": 200,
                "m3": 2.0,
                "placed_x_cm": 120,
                "placed_y_cm": 0,
                "orient_L_cm": 100,
                "orient_W_cm": 80,
            },
        ]
    )

    kpi_df = build_container_kpi_rows(placements_df)
    binary = build_excel_report(placements_df, kpi_df)
    wb = load_workbook(filename=BytesIO(binary))

    assert wb.sheetnames == ["Summary", "Placements", "Layout", "ContainerKPI"]
    assert wb["Summary"]["A1"].value == "container_label"
    assert wb["Summary"]["C2"].value == 2
    assert wb["Layout"]["A1"].value == "40HC ① Layout"
    assert wb["ContainerKPI"]["A1"].value == "container_label"
    assert wb["ContainerKPI"]["C2"].value == 3.2
    assert wb["ContainerKPI"]["F2"].value == 200
