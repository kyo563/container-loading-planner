from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def _to_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in cols:
        out[col] = pd.to_numeric(out[col], errors="coerce")
    return out


def _build_summary_sheet(wb: Workbook, placements_df: pd.DataFrame) -> pd.DataFrame:
    ws = wb.create_sheet("Summary")
    if placements_df.empty:
        headers = ["container_label", "container_type", "pieces", "total_weight_kg", "total_m3"]
        ws.append(headers)
        return pd.DataFrame(columns=headers)

    summary_df = (
        placements_df.groupby(["container_label", "container_type"], as_index=False)
        .agg(pieces=("cargo_piece_id", "count"), total_weight_kg=("weight_kg", "sum"), total_m3=("m3", "sum"))
        .sort_values(["container_type", "container_label"])
    )

    ws.append(list(summary_df.columns))
    for row in summary_df.itertuples(index=False, name=None):
        ws.append(list(row))
    _auto_width(ws)
    return summary_df


def _build_placements_sheet(wb: Workbook, placements_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Placements")
    if placements_df.empty:
        ws.append(["No placement data"])
        return

    ws.append(list(placements_df.columns))
    for row in placements_df.itertuples(index=False, name=None):
        ws.append(list(row))
    _auto_width(ws)


def _draw_container_layout(ws, start_row: int, title: str, container_df: pd.DataFrame) -> int:
    ws.cell(row=start_row, column=1, value=title)
    base_row = start_row + 2
    base_col = 2

    numeric_df = _to_numeric(container_df, ["placed_x_cm", "placed_y_cm", "orient_L_cm", "orient_W_cm"])

    max_x = (numeric_df["placed_x_cm"] + numeric_df["orient_L_cm"]).max()
    max_y = (numeric_df["placed_y_cm"] + numeric_df["orient_W_cm"]).max()
    if pd.isna(max_x) or pd.isna(max_y):
        ws.cell(row=base_row, column=1, value="配置座標データなし")
        return base_row + 2

    scale_cm = 20
    grid_w = int(max(1, min(80, (max_x // scale_cm) + 2)))
    grid_h = int(max(1, min(80, (max_y // scale_cm) + 2)))

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(fill_type="solid", start_color="A7D3F4", end_color="A7D3F4")

    for gy in range(grid_h):
        for gx in range(grid_w):
            cell = ws.cell(row=base_row + gy, column=base_col + gx)
            cell.border = border

    for row in numeric_df.itertuples(index=False):
        x0 = int(max(0, row.placed_x_cm // scale_cm))
        y0 = int(max(0, row.placed_y_cm // scale_cm))
        x1 = int(max(x0 + 1, (row.placed_x_cm + row.orient_L_cm) // scale_cm + 1))
        y1 = int(max(y0 + 1, (row.placed_y_cm + row.orient_W_cm) // scale_cm + 1))

        for gy in range(y0, min(y1, grid_h)):
            for gx in range(x0, min(x1, grid_w)):
                ws.cell(row=base_row + gy, column=base_col + gx).fill = fill

    legend_row = base_row + grid_h + 1
    ws.cell(row=legend_row, column=1, value=f"Scale: 1セル={scale_cm}cm")
    return legend_row + 2


def _build_layout_sheet(wb: Workbook, placements_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Layout")
    if placements_df.empty:
        ws.append(["No placement data"])
        return

    current_row = 1
    grouped = placements_df.sort_values(["container_type", "container_index"]).groupby("container_label")
    for container_label, group in grouped:
        current_row = _draw_container_layout(ws, current_row, f"{container_label} Layout", group)

    ws.column_dimensions["A"].width = 28


def build_excel_report(placements_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, placements_df)
    _build_placements_sheet(wb, placements_df)
    _build_layout_sheet(wb, placements_df)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
